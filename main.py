import openpyxl
import nltk
import fileinput
from collections import defaultdict
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter.filedialog import askopenfile
from tkinter import messagebox
import json
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup
from openpyxl.formatting.rule import CellIsRule

green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
light_green_fill = PatternFill(start_color="FF99FF99", end_color="FF99FF99", fill_type="solid")
lighter_green_fill = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid")
gray_fill = PatternFill(start_color='A9A9A9', end_color='A9A9A9', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
open_red_fill = PatternFill(start_color="FFFF6666", end_color="FFFF6666", fill_type="solid")
lighter_red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")


my_w = tk.Tk()
my_w.geometry("440x275")
my_w.title('FMCACBH')
my_font1=('Arial', 18, 'bold')
my_font2=('Arial', 10)

img = PhotoImage(file="stars.png")
l2 = tk.Label(my_w, image=img)
l2.pack(fill=X, pady=5)

l1 = tk.Label(my_w,text='SELECT YOUR .html FILE',width=30,font=my_font1)
l1.pack(fill=X, pady=5)

b1 = tk.Button(my_w, text='For Squad Developing Analysis', width=20, command=lambda:squadAnalysis())
b1.pack(fill=X, pady=5)

b2 = tk.Button(my_w, text='For Moneyball Att. Analysis', width=20, command=lambda:moneyballAttMentality())
b2.pack(fill=X, pady=5)

b3 = tk.Button(my_w, text='For Moneyball Sta. Analysis', width=20, command=lambda:moneyballStaMentality())
b3.pack(fill=X, pady=5)

b4 = tk.Button(my_w, text='For Coach and Staff Analysis', width=20, command=lambda:coachAnalysis())
b4.pack(fill=X, pady=5)

l3 = tk.Label(my_w,text='made by kemoguBH',width=20,font=my_font2)
l3.pack(fill=X, pady=5)

my_w.columnconfigure(0, weight=1)
my_w.rowconfigure(0, weight=1)

def squadAnalysis() :
    file_path = filedialog.askopenfilename(filetypes=[("HTML Dosyaları", "*.html")])

    with open(file_path, 'r', encoding='utf-8') as file:
        html = file.read()
        soup = BeautifulSoup(html, 'html.parser')

        table = soup.find('table')

        headers = table.find_all('th')
        header_names = [header.text.strip() for header in headers]

        rows = table.find_all('tr')

        data1 = []

        for row in rows[1:]:
            cells = row.find_all('td')
            row_data = [cell.text.strip() for cell in cells]

            if len(row_data) == len(header_names):
                playerAttsSeason1 = [int(att) for att in row_data[1:48]]
                playerDictSeason1 = {
                    "Name":row_data[0],
                    "Atts Season 1":playerAttsSeason1
                }
                data1.append(playerDictSeason1)

    file_path = filedialog.askopenfilename(filetypes=[("HTML Dosyaları", "*.html")])

    with open(file_path, 'r', encoding='utf-8') as file:
        html = file.read()
        soup = BeautifulSoup(html, 'html.parser')

        table = soup.find('table')

        headers = table.find_all('th')
        header_names = [header.text.strip() for header in headers]

        rows = table.find_all('tr')

        data2 = []

        for row in rows[1:]:
            cells = row.find_all('td')
            row_data = [cell.text.strip() for cell in cells]

            if len(row_data) == len(header_names):
                playerAttsSeason2 = [int(att) for att in row_data[1:48]]
                playerDictSeason2 = {
                    "Name":row_data[0],
                    "Atts Season 2":playerAttsSeason2
                }
                data2.append(playerDictSeason2)

    new_workbook = Workbook()
    new_sheet = new_workbook.active

    header = ["Name",
              "Acc", "Aer", "Agg", "Agi", "Ant", "Bal", "Bra",
              "Cmd", "Com",
              "Cmp", "Cnt", "Cor",
              "Cro", "Dec", "Det", "Dri",
              "Ecc", "Fin", "Fir", "Fla",
              "Fre", "Han", "Hea", "Jum", "Kic",
              "Ldr", "Lon", "L Th", "Mar", "Nat",
              "OtB", "1v1", "Pac", "Pas", "Pen",
              "Pos", "Pun", "Ref", "TRO", "Sta",
              "Str", "Tck", "Tea", "Tec", "Thr",
              "Vis", "Wor", "Total Develop"]
    new_sheet.append(header)

    dataFinal = []
    for player1 in data1:
        for player2 in data2:
            if player1["Name"] == player2["Name"]:
                atts_season1 = player1["Atts Season 1"]
                atts_season2 = player2["Atts Season 2"]
                atts_diff_list = [att2 - att1 for att1, att2 in zip(atts_season1, atts_season2)]
                atts_diff_sum = 0
                for platt in atts_diff_list:
                    atts_diff_sum += platt
                playerDictFinal = {
                    "Name" : player1["Name"],
                    "Atts Season Final" : player2["Atts Season 2"],
                    "Dif" : atts_diff_list,
                    "DifSum" : atts_diff_sum
                }
                dataFinal.append(playerDictFinal)

    for playerDictFinal in dataFinal:
        row_data = [playerDictFinal["Name"]] + playerDictFinal["Atts Season Final"] + [playerDictFinal["DifSum"]]
        print(row_data)
        new_sheet.append(row_data)
        for i, diff in enumerate(playerDictFinal["Dif"]):
            if diff == -1:
                cell = new_sheet.cell(row=new_sheet.max_row, column=i + 2)
                cell.fill = lighter_red_fill
            elif diff == -2:
                cell = new_sheet.cell(row=new_sheet.max_row, column=i + 2)
                cell.fill = open_red_fill
            elif diff <= -3:
                cell = new_sheet.cell(row=new_sheet.max_row, column=i + 2)
                cell.fill = red_fill
            elif diff == 1:
                cell = new_sheet.cell(row=new_sheet.max_row, column=i + 2)
                cell.fill = lighter_green_fill
            elif diff == 2:
                cell = new_sheet.cell(row=new_sheet.max_row, column=i + 2)
                cell.fill = light_green_fill
            elif diff >= 3:
                cell = new_sheet.cell(row=new_sheet.max_row, column=i + 2)
                cell.fill = green_fill

    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dosyaları", "*.xlsx")])
    new_workbook.save(save_path)

def moneyballAttMentality():
    def convert_value(value_str):
        value_str = str(value_str)
        if 'Sale' in value_str:
            return 999999999999999
        elif '-' in value_str:
            parts = value_str.split('-')
            right_part = parts[1].strip()
            if 'M' in right_part:
                right_part = right_part.replace('M', '').strip()
                if '€' in right_part:
                    right_part = right_part.replace('€', '').strip()
                return int(float(right_part) * 1000000)
            elif 'K' in right_part:
                right_part = right_part.replace('K', '').strip()
                if '€' in right_part:
                    right_part = right_part.replace('€', '').strip()
                return int(float(right_part) * 1000)
            else:
                if '€' in right_part:
                    right_part = right_part.replace('€', '').strip()
                return int(right_part)
        elif 'M' in value_str:
            return int(float(value_str.replace('€', '').replace('M', '')) * 1000000)
        elif 'K' in value_str:
            return int(float(value_str.replace('€', '').replace('K', '')) * 1000)
        else:
            return int(value_str.replace('€', ''))

    with open('atts.json', 'r') as json_file:
        json_data = json.load(json_file)

    gDefSum = sum(json_data['gDef'].values())
    skDefSum = sum(json_data['skDef'].values())
    skSupSum = sum(json_data['skSup'].values())
    skAtSum = sum(json_data['skAt'].values())
    fbDefSum = sum(json_data['fbDef'].values())
    fbSupSum = sum(json_data['fbSup'].values())
    fbAtSum = sum(json_data['fbAt'].values())
    wbDefSum = sum(json_data['wbDef'].values())
    wbSupSum = sum(json_data['wbSup'].values())
    wbAtSum = sum(json_data['wbAt'].values())
    cwbSupSum = sum(json_data['cwbSup'].values())
    cwbAtSum = sum(json_data['cwbAt'].values())
    iwbDefSum = sum(json_data['iwbDef'].values())
    iwbSupSum = sum(json_data['iwbSup'].values())
    iwbAtSum = sum(json_data['iwbAt'].values())
    nnfbDefSum = sum(json_data['nnfbDef'].values())
    nncbDefSum = sum(json_data['nncbDef'].values())
    nncbStSum = sum(json_data['nncbSt'].values())
    nncbCoSum = sum(json_data['nncbCo'].values())
    cdDefSum = sum(json_data['cdDef'].values())
    cdStSum = sum(json_data['cdSt'].values())
    cdCoSum = sum(json_data['cdCo'].values())
    bpdDefSum = sum(json_data['bpdDef'].values())
    bpdStSum = sum(json_data['bpdSt'].values())
    bpdCoSum = sum(json_data['bpdCo'].values())
    lSupSum = sum(json_data['lSup'].values())
    lAtSum = sum(json_data['lAt'].values())
    wcbDefSum = sum(json_data['wcbDef'].values())
    wcbSupSum = sum(json_data['wcbSup'].values())
    wcbAtSum = sum(json_data['wcbAt'].values())
    dmDefSum = sum(json_data['dmDef'].values())
    dmSupSum = sum(json_data['dmSup'].values())
    aDefSum = sum(json_data['aDef'].values())
    hbDefSum = sum(json_data['hbDef'].values())
    regSupSum = sum(json_data['regSup'].values())
    volSupSum = sum(json_data['volSup'].values())
    volAtSum = sum(json_data['volAt'].values())
    bwmDefSum = sum(json_data['bwmDef'].values())
    bwmSupSum = sum(json_data['bwmSup'].values())
    dlpDefSum = sum(json_data['dlpDef'].values())
    dlpSupSum = sum(json_data['dlpSup'].values())
    rpSupSum = sum(json_data['rpSup'].values())
    apSupSum = sum(json_data['apSup'].values())
    apAtSum = sum(json_data['apAt'].values())
    cmDefSum = sum(json_data['cmDef'].values())
    cmSupSum = sum(json_data['cmSup'].values())
    cmAtSum = sum(json_data['cmAt'].values())
    b2bSupSum = sum(json_data['b2bSup'].values())
    carSupSum = sum(json_data['carSup'].values())
    mezSupSum = sum(json_data['mezSup'].values())
    mezAtSum = sum(json_data['mezAt'].values())
    wSupSum = sum(json_data['wSup'].values())
    wAtSum = sum(json_data['wAt'].values())
    dwDefSum = sum(json_data['dwDef'].values())
    dwSupSum = sum(json_data['dwSup'].values())
    wmDefSum = sum(json_data['wmDef'].values())
    wmSupSum = sum(json_data['wmSup'].values())
    wmAtSum = sum(json_data['wmAt'].values())
    wpSupSum = sum(json_data['wpSup'].values())
    wpAtSum = sum(json_data['wpAt'].values())
    iwSupSum = sum(json_data['iwSup'].values())
    iwAtSum = sum(json_data['iwAt'].values())
    treqAtSum = sum(json_data['treqAt'].values())
    raumAtSum = sum(json_data['raumAt'].values())
    wtfSupSum = sum(json_data['wtfSup'].values())
    wtfAtSum = sum(json_data['wtfAt'].values())
    ifSupSum = sum(json_data['ifSup'].values())
    ifAtSum = sum(json_data['ifAt'].values())
    amSupSum = sum(json_data['amSup'].values())
    amAtSum = sum(json_data['amAt'].values())
    engSupSum = sum(json_data['engSup'].values())
    ssAtSum = sum(json_data['ssAt'].values())
    f9SupSum = sum(json_data['f9Sup'].values())
    tfSupSum = sum(json_data['tfSup'].values())
    tfatSum = sum(json_data['tfat'].values())
    dlfSupSum = sum(json_data['dlfSup'].values())
    dlfAtSum = sum(json_data['dlfAt'].values())
    pfDefSum = sum(json_data['pfDef'].values())
    pfSupSum = sum(json_data['pfSup'].values())
    pfAtSum = sum(json_data['pfAt'].values())
    cfSupSum = sum(json_data['cfSup'].values())
    cfAtSum = sum(json_data['cfAt'].values())
    pAtSum = sum(json_data['pAt'].values())
    afAtSum = sum(json_data['afAt'].values())

    file_path = filedialog.askopenfilename(filetypes=[("HTML Dosyaları", "*.html")])

    with open(file_path, 'r', encoding='utf-8') as file:
        html = file.read()
        soup = BeautifulSoup(html, 'html.parser')

        table = soup.find('table')

        headers = table.find_all('th')
        header_names = [header.text.strip() for header in headers]

        rows = table.find_all('tr')

        data = []

        for row in rows[1:]:
            cells = row.find_all('td')
            row_data = [cell.text.strip() for cell in cells]

            if len(row_data) == len(header_names):
                playerName = row_data[0]
                player_att = [int(att) for att in row_data[1:48]]
                player_value = convert_value(row_data[48])
                print(player_value)
                player_age = row_data[49]
                print(player_age)
                player_club = row_data[50]
                print(player_club)
                player_personality = row_data[51]
                print(player_personality)
                player_media_description = row_data[52]
                print(player_media_description)

                gDef = sum(att_value * json_data['gDef'][att_name] for att_name, att_value in
                           zip(json_data['gDef'].keys(), player_att))
                skDef = sum(att_value * json_data['skDef'][att_name] for att_name, att_value in
                            zip(json_data['skDef'].keys(), player_att))
                skSup = sum(att_value * json_data['skSup'][att_name] for att_name, att_value in
                            zip(json_data['skSup'].keys(), player_att))
                skAt = sum(att_value * json_data['skAt'][att_name] for att_name, att_value in
                           zip(json_data['skAt'].keys(), player_att))
                fbDef = sum(att_value * json_data['fbDef'][att_name] for att_name, att_value in
                            zip(json_data['fbDef'].keys(), player_att))
                fbSup = sum(att_value * json_data['fbSup'][att_name] for att_name, att_value in
                            zip(json_data['fbSup'].keys(), player_att))
                fbAt = sum(att_value * json_data['fbAt'][att_name] for att_name, att_value in
                           zip(json_data['fbAt'].keys(), player_att))
                wbDef = sum(att_value * json_data['wbDef'][att_name] for att_name, att_value in
                            zip(json_data['wbDef'].keys(), player_att))
                wbSup = sum(att_value * json_data['wbSup'][att_name] for att_name, att_value in
                            zip(json_data['wbSup'].keys(), player_att))
                wbAt = sum(att_value * json_data['wbAt'][att_name] for att_name, att_value in
                           zip(json_data['wbAt'].keys(), player_att))
                cwbSup = sum(att_value * json_data['cwbSup'][att_name] for att_name, att_value in
                             zip(json_data['cwbSup'].keys(), player_att))
                cwbAt = sum(att_value * json_data['cwbAt'][att_name] for att_name, att_value in
                            zip(json_data['cwbAt'].keys(), player_att))
                iwbDef = sum(att_value * json_data['iwbDef'][att_name] for att_name, att_value in
                             zip(json_data['iwbDef'].keys(), player_att))
                iwbSup = sum(att_value * json_data['iwbSup'][att_name] for att_name, att_value in
                             zip(json_data['iwbSup'].keys(), player_att))
                iwbAt = sum(att_value * json_data['iwbAt'][att_name] for att_name, att_value in
                            zip(json_data['iwbAt'].keys(), player_att))
                nnfbDef = sum(att_value * json_data['nnfbDef'][att_name] for att_name, att_value in
                              zip(json_data['nnfbDef'].keys(), player_att))
                nncbDef = sum(att_value * json_data['nncbDef'][att_name] for att_name, att_value in
                              zip(json_data['nncbDef'].keys(), player_att))
                nncbSt = sum(att_value * json_data['nncbSt'][att_name] for att_name, att_value in
                             zip(json_data['nncbSt'].keys(), player_att))
                nncbCo = sum(att_value * json_data['nncbCo'][att_name] for att_name, att_value in
                             zip(json_data['nncbCo'].keys(), player_att))
                cdDef = sum(att_value * json_data['cdDef'][att_name] for att_name, att_value in
                            zip(json_data['cdDef'].keys(), player_att))
                cdSt = sum(att_value * json_data['cdSt'][att_name] for att_name, att_value in
                           zip(json_data['cdSt'].keys(), player_att))
                cdCo = sum(att_value * json_data['cdCo'][att_name] for att_name, att_value in
                           zip(json_data['cdCo'].keys(), player_att))
                bpdDef = sum(att_value * json_data['bpdDef'][att_name] for att_name, att_value in
                             zip(json_data['bpdDef'].keys(), player_att))
                bpdSt = sum(att_value * json_data['bpdSt'][att_name] for att_name, att_value in
                            zip(json_data['bpdSt'].keys(), player_att))
                bpdCo = sum(att_value * json_data['bpdCo'][att_name] for att_name, att_value in
                            zip(json_data['bpdCo'].keys(), player_att))
                lSup = sum(att_value * json_data['lSup'][att_name] for att_name, att_value in
                           zip(json_data['lSup'].keys(), player_att))
                lAt = sum(att_value * json_data['lAt'][att_name] for att_name, att_value in
                          zip(json_data['lAt'].keys(), player_att))
                wcbDef = sum(att_value * json_data['wcbDef'][att_name] for att_name, att_value in
                             zip(json_data['wcbDef'].keys(), player_att))
                wcbSup = sum(att_value * json_data['wcbSup'][att_name] for att_name, att_value in
                             zip(json_data['wcbSup'].keys(), player_att))
                wcbAt = sum(att_value * json_data['wcbAt'][att_name] for att_name, att_value in
                            zip(json_data['wcbAt'].keys(), player_att))
                dmDef = sum(att_value * json_data['dmDef'][att_name] for att_name, att_value in
                            zip(json_data['dmDef'].keys(), player_att))
                dmSup = sum(att_value * json_data['dmSup'][att_name] for att_name, att_value in
                            zip(json_data['dmSup'].keys(), player_att))
                aDef = sum(att_value * json_data['aDef'][att_name] for att_name, att_value in
                           zip(json_data['aDef'].keys(), player_att))
                hbDef = sum(att_value * json_data['hbDef'][att_name] for att_name, att_value in
                            zip(json_data['hbDef'].keys(), player_att))
                regSup = sum(att_value * json_data['regSup'][att_name] for att_name, att_value in
                             zip(json_data['regSup'].keys(), player_att))
                volSup = sum(att_value * json_data['volSup'][att_name] for att_name, att_value in
                             zip(json_data['volSup'].keys(), player_att))
                volAt = sum(att_value * json_data['volAt'][att_name] for att_name, att_value in
                            zip(json_data['volAt'].keys(), player_att))
                bwmDef = sum(att_value * json_data['bwmDef'][att_name] for att_name, att_value in
                             zip(json_data['bwmDef'].keys(), player_att))
                bwmSup = sum(att_value * json_data['bwmSup'][att_name] for att_name, att_value in
                             zip(json_data['bwmSup'].keys(), player_att))
                dlpDef = sum(att_value * json_data['dlpDef'][att_name] for att_name, att_value in
                             zip(json_data['dlpDef'].keys(), player_att))
                dlpSup = sum(att_value * json_data['dlpSup'][att_name] for att_name, att_value in
                             zip(json_data['dlpSup'].keys(), player_att))
                rpSup = sum(att_value * json_data['rpSup'][att_name] for att_name, att_value in
                            zip(json_data['rpSup'].keys(), player_att))
                apSup = sum(att_value * json_data['apSup'][att_name] for att_name, att_value in
                            zip(json_data['apSup'].keys(), player_att))
                apAt = sum(att_value * json_data['apAt'][att_name] for att_name, att_value in
                           zip(json_data['apAt'].keys(), player_att))
                cmDef = sum(att_value * json_data['cmDef'][att_name] for att_name, att_value in
                            zip(json_data['cmDef'].keys(), player_att))
                cmSup = sum(att_value * json_data['cmSup'][att_name] for att_name, att_value in
                            zip(json_data['cmSup'].keys(), player_att))
                cmAt = sum(att_value * json_data['cmAt'][att_name] for att_name, att_value in
                           zip(json_data['cmAt'].keys(), player_att))
                b2bSup = sum(att_value * json_data['b2bSup'][att_name] for att_name, att_value in
                             zip(json_data['b2bSup'].keys(), player_att))
                carSup = sum(att_value * json_data['carSup'][att_name] for att_name, att_value in
                             zip(json_data['carSup'].keys(), player_att))
                mezSup = sum(att_value * json_data['mezSup'][att_name] for att_name, att_value in
                             zip(json_data['mezSup'].keys(), player_att))
                mezAt = sum(att_value * json_data['mezAt'][att_name] for att_name, att_value in
                            zip(json_data['mezAt'].keys(), player_att))
                wSup = sum(att_value * json_data['wSup'][att_name] for att_name, att_value in
                           zip(json_data['wSup'].keys(), player_att))
                wAt = sum(att_value * json_data['wAt'][att_name] for att_name, att_value in
                          zip(json_data['wAt'].keys(), player_att))
                dwDef = sum(att_value * json_data['dwDef'][att_name] for att_name, att_value in
                            zip(json_data['dwDef'].keys(), player_att))
                dwSup = sum(att_value * json_data['dwSup'][att_name] for att_name, att_value in
                            zip(json_data['dwSup'].keys(), player_att))
                wmDef = sum(att_value * json_data['wmDef'][att_name] for att_name, att_value in
                            zip(json_data['wmDef'].keys(), player_att))
                wmSup = sum(att_value * json_data['wmSup'][att_name] for att_name, att_value in
                            zip(json_data['wmSup'].keys(), player_att))
                wmAt = sum(att_value * json_data['wmAt'][att_name] for att_name, att_value in
                           zip(json_data['wmAt'].keys(), player_att))
                wpSup = sum(att_value * json_data['wpSup'][att_name] for att_name, att_value in
                            zip(json_data['wpSup'].keys(), player_att))
                wpAt = sum(att_value * json_data['wpAt'][att_name] for att_name, att_value in
                           zip(json_data['wpAt'].keys(), player_att))
                iwSup = sum(att_value * json_data['iwSup'][att_name] for att_name, att_value in
                            zip(json_data['iwSup'].keys(), player_att))
                iwAt = sum(att_value * json_data['iwAt'][att_name] for att_name, att_value in
                           zip(json_data['iwAt'].keys(), player_att))
                treqAt = sum(att_value * json_data['treqAt'][att_name] for att_name, att_value in
                             zip(json_data['treqAt'].keys(), player_att))
                raumAt = sum(att_value * json_data['raumAt'][att_name] for att_name, att_value in
                             zip(json_data['raumAt'].keys(), player_att))
                wtfSup = sum(att_value * json_data['wtfSup'][att_name] for att_name, att_value in
                             zip(json_data['wtfSup'].keys(), player_att))
                wtfAt = sum(att_value * json_data['wtfAt'][att_name] for att_name, att_value in
                            zip(json_data['wtfAt'].keys(), player_att))
                ifSup = sum(att_value * json_data['ifSup'][att_name] for att_name, att_value in
                            zip(json_data['ifSup'].keys(), player_att))
                ifAt = sum(att_value * json_data['ifAt'][att_name] for att_name, att_value in
                           zip(json_data['ifAt'].keys(), player_att))
                amSup = sum(att_value * json_data['amSup'][att_name] for att_name, att_value in
                            zip(json_data['amSup'].keys(), player_att))
                amAt = sum(att_value * json_data['amAt'][att_name] for att_name, att_value in
                           zip(json_data['amAt'].keys(), player_att))
                engSup = sum(att_value * json_data['engSup'][att_name] for att_name, att_value in
                             zip(json_data['engSup'].keys(), player_att))
                ssAt = sum(att_value * json_data['ssAt'][att_name] for att_name, att_value in
                           zip(json_data['ssAt'].keys(), player_att))
                f9Sup = sum(att_value * json_data['f9Sup'][att_name] for att_name, att_value in
                            zip(json_data['f9Sup'].keys(), player_att))
                tfSup = sum(att_value * json_data['tfSup'][att_name] for att_name, att_value in
                            zip(json_data['tfSup'].keys(), player_att))
                tfat = sum(att_value * json_data['tfat'][att_name] for att_name, att_value in
                           zip(json_data['tfat'].keys(), player_att))
                dlfSup = sum(att_value * json_data['dlfSup'][att_name] for att_name, att_value in
                             zip(json_data['dlfSup'].keys(), player_att))
                dlfAt = sum(att_value * json_data['dlfAt'][att_name] for att_name, att_value in
                            zip(json_data['dlfAt'].keys(), player_att))
                pfDef = sum(att_value * json_data['pfDef'][att_name] for att_name, att_value in
                            zip(json_data['pfDef'].keys(), player_att))
                pfSup = sum(att_value * json_data['pfSup'][att_name] for att_name, att_value in
                            zip(json_data['pfSup'].keys(), player_att))
                pfAt = sum(att_value * json_data['pfAt'][att_name] for att_name, att_value in
                           zip(json_data['pfAt'].keys(), player_att))
                cfSup = sum(att_value * json_data['cfSup'][att_name] for att_name, att_value in
                            zip(json_data['cfSup'].keys(), player_att))
                cfAt = sum(att_value * json_data['cfAt'][att_name] for att_name, att_value in
                           zip(json_data['cfAt'].keys(), player_att))
                pAt = sum(att_value * json_data['pAt'][att_name] for att_name, att_value in
                          zip(json_data['pAt'].keys(), player_att))
                afAt = sum(att_value * json_data['afAt'][att_name] for att_name, att_value in
                           zip(json_data['afAt'].keys(), player_att))

                playerDict = {
                    "Name": playerName,
                    "Att": player_att,
                    "G-DE": gDef,
                    "SK-DE": skDef,
                    "SK-SU": skSup,
                    "SK-AT": skAt,
                    "FB-DE": fbDef,
                    "FB-SU": fbSup,
                    "FB-AT": fbAt,
                    "WB-DE": wbDef,
                    "WB-SU": wbSup,
                    "WB-AT": wbAt,
                    "CWB-SU": cwbSup,
                    "CWB-AT": cwbAt,
                    "IWB-DE": iwbDef,
                    "IWB-SU": iwbSup,
                    "IWB-AT": iwbAt,
                    "NNFB-DE": nnfbDef,
                    "NNCB-DE": nncbDef,
                    "NNCB-ST": nncbSt,
                    "NNCB-CO": nncbCo,
                    "CD-DE": cdDef,
                    "CD-ST": cdSt,
                    "CD-CO": cdCo,
                    "BPD-DE": bpdDef,
                    "BPD-ST": bpdSt,
                    "BPD-CO": bpdCo,
                    "L-SU": lSup,
                    "L-AT": lAt,
                    "WCB-DE": wcbDef,
                    "WCB-SU": wcbSup,
                    "WCB-AT": wcbAt,
                    "DM-DE": dmDef,
                    "DM-SU": dmSup,
                    "A-DE": aDef,
                    "HB-DE": hbDef,
                    "REG-SU": regSup,
                    "VOL-SU": volSup,
                    "VOL-AT": volSup,
                    "BWM-DE": bwmDef,
                    "BWM-SU": bwmSup,
                    "DLP-DE": dlpDef,
                    "DLP-SU": dlpSup,
                    "RP-SU": rpSup,
                    "AP-SU": apSup,
                    "AP-AT": apAt,
                    "CM-DE": cmDef,
                    "CM-SU": cmSup,
                    "CM-AT": cmAt,
                    "B2B-SU": b2bSup,
                    "CAR-SU": carSup,
                    "MEZ-SU": mezSup,
                    "MEZ-AT": mezAt,
                    "W-SU": wSup,
                    "W-AT": wAt,
                    "DW-DE": dwDef,
                    "DW-SU": dwSup,
                    "WM-DE": wmDef,
                    "WM-SU": wmSup,
                    "WM-AT": wmAt,
                    "WP-SU": wpSup,
                    "WP-AT": wpAt,
                    "IW-SU": iwSup,
                    "IW-AT": iwAt,
                    "T-AT": treqAt,
                    "R-AT": raumAt,
                    "WTF-SU": wtfSup,
                    "WTF-AT": wtfAt,
                    "IF-SU": ifSup,
                    "IF-AT": ifAt,
                    "AM-SU": amSup,
                    "AM-AT": amAt,
                    "ENG-SU": engSup,
                    "SS-AT": ssAt,
                    "F9-SU": f9Sup,
                    "TF-SU": tfSup,
                    "TF-AT": tfat,
                    "DLF-SU": dlfSup,
                    "DLF-AT": dlfAt,
                    "PF-DE": pfDef,
                    "PF-SU": pfSup,
                    "PF-AT": pfAt,
                    "CF-SU": cfSup,
                    "CF-AT": cfAt,
                    "P-AT": pAt,
                    "AF-AT": afAt,
                    "Value": player_value,
                    "Age":player_age,
                    "Club":player_club,
                    "Personality" : player_personality,
                    "Media Description" : player_media_description
                }
                data.append(playerDict)

    new_workbook = Workbook()
    new_sheet = new_workbook.active

    new_sheet.append([
        "Name",
        "G-DE",
        "SK-DE",
        "SK-SU",
        "SK-AT",
        "FB-DE",
        "FB-SU",
        "FB-AT",
        "WB-DE",
        "WB-SU",
        "WB-AT",
        "CWB-SU",
        "CWB-AT",
        "IWB-DE",
        "IWB-SU",
        "IWB-AT",
        "NNFB-DE",
        "NNCB-DE",
        "NNCB-ST",
        "NNCB-CO",
        "CD-DE",
        "CD-ST",
        "CD-CO",
        "BPD-DE",
        "BPD-ST",
        "BPD-Co",
        "L-SU",
        "L-AT",
        "WCB-DE",
        "WCB-SU",
        "WCB-AT",
        "DM-DE",
        "DM-SU",
        "A-DE",
        "HB-DE",
        "REG-SU",
        "VOL-SU",
        "VOL-AT",
        "BWM-DE",
        "BWM-SU",
        "DLP-DE",
        "DLP-SU",
        "RP-SU",
        "AP-SU",
        "AP-AT",
        "CM-DE",
        "CM-SU",
        "CM-AT",
        "B2B-SU",
        "CAR-SU",
        "MEZ-SU",
        "MEZ-AT",
        "W-SU",
        "W-AT",
        "DW-DE",
        "DW-SU",
        "WM-DE",
        "WM-SU",
        "WM-AT",
        "WP-SU",
        "WP-AT",
        "IW-SU",
        "IW-AT",
        "T-AT",
        "R-AT",
        "WTF-SU",
        "WTF-AT",
        "IF-SU",
        "IF-AT",
        "AM-SU",
        "AM-AT",
        "ENG-SU",
        "SS-AT",
        "F9-SU",
        "TF-SU",
        "TF-AT",
        "DLF-SU",
        "DLF-AT",
        "PF-DE",
        "PF-SU",
        "PF-AT",
        "CF-SU",
        "CF-AT",
        "P-AT",
        "AF-AT",
        "Value",
        "Age",
        "Club",
        "Personality",
        "Media Description",
        "Best Role",
        "Best Score",
        "Score/Value"
    ])

    for player in data:
        new_sheet.append([
            player["Name"],
            round(player['G-DE'] / gDefSum, 1),
            round(player['SK-DE'] / skDefSum, 1),
            round(player['SK-SU'] / skSupSum, 1),
            round(player['SK-AT'] / skAtSum, 1),
            round(player['FB-DE'] / fbDefSum, 1),
            round(player['FB-SU'] / fbSupSum, 1),
            round(player['FB-AT'] / fbAtSum, 1),
            round(player['WB-DE'] / wbDefSum, 1),
            round(player['WB-SU'] / wbSupSum, 1),
            round(player['WB-AT'] / wbAtSum, 1),
            round(player['CWB-SU'] / cwbSupSum, 1),
            round(player['CWB-AT'] / cwbAtSum, 1),
            round(player['IWB-DE'] / iwbDefSum, 1),
            round(player['IWB-SU'] / iwbSupSum, 1),
            round(player['IWB-AT'] / iwbAtSum, 1),
            round(player['NNFB-DE'] / nnfbDefSum, 1),
            round(player['NNCB-DE'] / nncbDefSum, 1),
            round(player['NNCB-ST'] / nncbStSum, 1),
            round(player['NNCB-CO'] / nncbCoSum, 1),
            round(player['CD-DE'] / cdDefSum, 1),
            round(player['CD-ST'] / cdStSum, 1),
            round(player['CD-CO'] / cdCoSum, 1),
            round(player['BPD-DE'] / bpdDefSum, 1),
            round(player['BPD-ST'] / bpdStSum, 1),
            round(player['BPD-CO'] / bpdCoSum, 1),
            round(player['L-SU'] / lSupSum, 1),
            round(player['L-AT'] / lAtSum, 1),
            round(player['WCB-DE'] / wcbDefSum, 1),
            round(player['WCB-SU'] / wcbSupSum, 1),
            round(player['WCB-AT'] / wcbAtSum, 1),
            round(player['DM-DE'] / dmDefSum, 1),
            round(player['DM-SU'] / dmSupSum, 1),
            round(player['A-DE'] / aDefSum, 1),
            round(player['HB-DE'] / hbDefSum, 1),
            round(player['REG-SU'] / regSupSum, 1),
            round(player['VOL-SU'] / volSupSum, 1),
            round(player['VOL-AT'] / volAtSum, 1),
            round(player['BWM-DE'] / bwmDefSum, 1),
            round(player['BWM-SU'] / bwmSupSum, 1),
            round(player['DLP-DE'] / dlpDefSum, 1),
            round(player['DLP-SU'] / dlpSupSum, 1),
            round(player['RP-SU'] / rpSupSum, 1),
            round(player['AP-SU'] / apSupSum, 1),
            round(player['AP-AT'] / apAtSum, 1),
            round(player['CM-DE'] / cmDefSum, 1),
            round(player['CM-SU'] / cmSupSum, 1),
            round(player['CM-AT'] / cmAtSum, 1),
            round(player['B2B-SU'] / b2bSupSum, 1),
            round(player['CAR-SU'] / carSupSum, 1),
            round(player['MEZ-SU'] / mezSupSum, 1),
            round(player['MEZ-AT'] / mezAtSum, 1),
            round(player['W-SU'] / wSupSum, 1),
            round(player['W-AT'] / wAtSum, 1),
            round(player['DW-DE'] / dwDefSum, 1),
            round(player['DW-SU'] / dwSupSum, 1),
            round(player['WM-DE'] / wmDefSum, 1),
            round(player['WM-SU'] / wmSupSum, 1),
            round(player['WM-AT'] / wmAtSum, 1),
            round(player['WP-SU'] / wpSupSum, 1),
            round(player['WP-AT'] / wpAtSum, 1),
            round(player['IW-SU'] / iwSupSum, 1),
            round(player['IW-AT'] / iwAtSum, 1),
            round(player['T-AT'] / treqAtSum, 1),
            round(player['R-AT'] / raumAtSum, 1),
            round(player['WTF-SU'] / wtfSupSum, 1),
            round(player['WTF-AT'] / wtfAtSum, 1),
            round(player['IF-SU'] / ifSupSum, 1),
            round(player['IF-AT'] / ifAtSum, 1),
            round(player['AM-SU'] / amSupSum, 1),
            round(player['AM-AT'] / amAtSum, 1),
            round(player['ENG-SU'] / engSupSum, 1),
            round(player['SS-AT'] / ssAtSum, 1),
            round(player['F9-SU'] / f9SupSum, 1),
            round(player['TF-SU'] / tfSupSum, 1),
            round(player['TF-AT'] / tfatSum, 1),
            round(player['DLF-SU'] / dlfSupSum, 1),
            round(player['DLF-AT'] / dlfAtSum, 1),
            round(player['PF-DE'] / pfDefSum, 1),
            round(player['PF-SU'] / pfSupSum, 1),
            round(player['PF-AT'] / pfAtSum, 1),
            round(player['CF-SU'] / cfSupSum, 1),
            round(player['CF-AT'] / cfAtSum, 1),
            round(player['P-AT'] / pAtSum, 1),
            round(player['AF-AT'] / afAtSum, 1),
            player["Value"],
            player["Age"],
            player["Club"],
            player["Personality"],
            player["Media Description"]
        ])

    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dosyaları", "*.xlsx")])

    new_workbook.save(save_path)

    new_workbook = openpyxl.load_workbook(save_path)

    sheet = new_workbook.active

    start_column = 2
    end_column = 85


    for row_number, row in enumerate(sheet.iter_rows(min_row=2, min_col=start_column, max_col=end_column), start=2):
        max_value = None
        max_column = None
        for cell in row:
            try:
                cell_value = cell.value
                if cell_value <= 5 :
                    cell.fill = gray_fill
                elif cell_value > 5 and cell_value <= 10 :
                    pass
                elif cell_value > 10 and cell_value <= 15 :
                    cell.fill = yellow_fill
                elif cell_value > 15 and cell_value <= 20:
                    cell.fill = green_fill
            except ValueError:
                continue

            if max_value is None or cell_value > max_value:
                max_value = cell_value
                max_column = cell.column
            first_element_of_max_column = None
            for row in sheet.iter_rows(min_row=1,max_row=1, min_col=max_column, max_col=max_column):
                for cell in row:
                    first_element_of_max_column = cell.value
                    break
                break
        sheet.cell(row=row_number, column=91).value = first_element_of_max_column
        sheet.cell(row = row_number ,column=92).value= max_value
        if max_value <= 5:
            sheet.cell(row = row_number ,column=92).fill = gray_fill
        elif max_value > 5 and max_value <= 10:
            pass
        elif max_value > 10 and max_value <= 15:
            sheet.cell(row = row_number ,column=92).fill = yellow_fill
        elif max_value > 15 and max_value <= 20:
            sheet.cell(row = row_number ,column=92).fill = green_fill
        if sheet.cell(row=row_number, column=86).value == 0:
            sheet.cell(row=row_number, column=86).value = 1
        sheet.cell(row=row_number, column=93).value = max_value / sheet.cell(row=row_number, column=86).value

    new_workbook.save(save_path)

def coachAnalysis() :
    with open('coachAtts.json', 'r') as json_file:
        json_data = json.load(json_file)

    managerSum = sum(json_data['manager'].values())
    assistantManagerSum = sum(json_data['assistantManager'].values())
    headOfYouthDevelopmentSum = sum(json_data['headOfYouthDevelopment'].values())
    shotStoppingSum = sum(json_data['goalkeepingCoach-ShotStopping'].values())
    handlingandDistributionSum = sum(json_data['goalkeepingCoach-HandlingandDistribution'].values())
    fitnessCoachStrengthSum = sum(json_data['fitnessCoach-Strength'].values())
    fitnessCoachQuicknessSum = sum(json_data['fitnessCoach-Quickness'].values())
    defendingTechnicalSum = sum(json_data['defending-Technical'].values())
    defendingTacticalSum = sum(json_data['defending-Tactical'].values())
    possessionTechnicalSum = sum(json_data['possession-Technical'].values())
    possessionTacticalSum = sum(json_data['possession-Tactical'].values())
    attackingTechnicalSum = sum(json_data['attacking-Technical'].values())
    attackingTacticalSum = sum(json_data['attacking-Tactical'].values())
    headPerformanceAnalystSum = sum(json_data['headPerformanceAnalyst'].values())
    performanceAnalystSum = sum(json_data['performanceAnalyst'].values())
    directorOfFootballSum = sum(json_data['directorOfFootball'].values())
    technicalDirectorSum = sum(json_data['technicalDirector'].values())
    chiefScoutSum = sum(json_data['chiefScout'].values())
    scoutSum = sum(json_data['scout'].values())
    recruitmentAnalystSum = sum(json_data['recruitmentAnalyst'].values())
    loanManagerSum = sum(json_data['loanManager'].values())
    headPhysioSum = sum(json_data['headPhysio'].values())
    headOfSportsScienceSum = sum(json_data['headOfSportsScience'].values())
    physioSum = sum(json_data['physio'].values())
    sportScientistSum = sum(json_data['sportScientist'].values())

    file_path = filedialog.askopenfilename(filetypes=[("HTML Dosyaları", "*.html")])

    with open(file_path, 'r', encoding='utf-8') as file:
        html = file.read()
        soup = BeautifulSoup(html, 'html.parser')

        table = soup.find('table')

        headers = table.find_all('th')
        header_names = [header.text.strip() for header in headers]

        rows = table.find_all('tr')

        data = []

        for row in rows[1:]:
            cells = row.find_all('td')
            row_data = [cell.text.strip() for cell in cells]

            if len(row_data) == len(header_names):
                coachAtts = [int(att) for att in row_data[7:]]  # Convert string values to integers
                manager = round(sum(att_value * json_data['manager'][att_name] for att_name, att_value in
                              zip(json_data['manager'].keys(), coachAtts))/managerSum,1)
                assistantManager = round(sum(att_value * json_data['assistantManager'][att_name] for att_name, att_value in
                              zip(json_data['assistantManager'].keys(), coachAtts))/assistantManagerSum,1)
                headOfYouthDevelopment = round(sum(att_value * json_data['headOfYouthDevelopment'][att_name] for att_name, att_value in
                              zip(json_data['headOfYouthDevelopment'].keys(), coachAtts))/headOfYouthDevelopmentSum,1)
                goalkeepingCoachShotStopping = sum(att_value * json_data['goalkeepingCoach-ShotStopping'][att_name] for att_name, att_value in
                              zip(json_data['goalkeepingCoach-ShotStopping'].keys(), coachAtts))/shotStoppingSum
                goalkeepingCoachHandlingandDistribution = sum(att_value * json_data['goalkeepingCoach-HandlingandDistribution'][att_name] for att_name, att_value in
                              zip(json_data['goalkeepingCoach-HandlingandDistribution'].keys(), coachAtts))/handlingandDistributionSum
                fitnessCoachStrength = sum(att_value * json_data['fitnessCoach-Strength'][att_name] for att_name, att_value in
                              zip(json_data['fitnessCoach-Strength'].keys(), coachAtts))/fitnessCoachStrengthSum
                fitnessCoachQuickness = sum(att_value * json_data['fitnessCoach-Quickness'][att_name] for att_name, att_value in
                              zip(json_data['fitnessCoach-Quickness'].keys(), coachAtts))/fitnessCoachQuicknessSum
                defendingTechnical = sum(att_value * json_data['defending-Technical'][att_name] for att_name, att_value in
                              zip(json_data['defending-Technical'].keys(), coachAtts))/defendingTechnicalSum
                defendingTactical = sum(att_value * json_data['defending-Tactical'][att_name] for att_name, att_value in
                              zip(json_data['defending-Tactical'].keys(), coachAtts))/defendingTacticalSum
                possessionTechnical = sum(att_value * json_data['possession-Technical'][att_name] for att_name, att_value in
                              zip(json_data['possession-Technical'].keys(), coachAtts))/possessionTechnicalSum
                possessionTactical = sum(att_value * json_data['possession-Tactical'][att_name] for att_name, att_value in
                              zip(json_data['possession-Tactical'].keys(), coachAtts))/possessionTacticalSum
                attackingTechnical = sum(att_value * json_data['attacking-Technical'][att_name] for att_name, att_value in
                              zip(json_data['attacking-Technical'].keys(), coachAtts))/attackingTechnicalSum
                attackingTactical = sum(att_value * json_data['attacking-Tactical'][att_name] for att_name, att_value in
                              zip(json_data['attacking-Tactical'].keys(), coachAtts))/attackingTacticalSum
                headPerformanceAnalyst = sum(att_value * json_data['headPerformanceAnalyst'][att_name] for att_name, att_value in
                              zip(json_data['headPerformanceAnalyst'].keys(), coachAtts))/headPerformanceAnalystSum
                performanceAnalyst = sum(att_value * json_data['performanceAnalyst'][att_name] for att_name, att_value in
                              zip(json_data['performanceAnalyst'].keys(), coachAtts))/performanceAnalystSum
                directorOfFootball = sum(att_value * json_data['directorOfFootball'][att_name] for att_name, att_value in
                              zip(json_data['directorOfFootball'].keys(), coachAtts))/directorOfFootballSum
                technicalDirector = sum(att_value * json_data['technicalDirector'][att_name] for att_name, att_value in
                              zip(json_data['technicalDirector'].keys(), coachAtts))/technicalDirectorSum
                chiefScout = sum(att_value * json_data['chiefScout'][att_name] for att_name, att_value in
                              zip(json_data['chiefScout'].keys(), coachAtts))/chiefScoutSum
                scout = sum(att_value * json_data['scout'][att_name] for att_name, att_value in
                              zip(json_data['scout'].keys(), coachAtts))/scoutSum
                recruitmentAnalyst = sum(att_value * json_data['recruitmentAnalyst'][att_name] for att_name, att_value in
                              zip(json_data['recruitmentAnalyst'].keys(), coachAtts))/recruitmentAnalystSum
                loanManager = sum(att_value * json_data['loanManager'][att_name] for att_name, att_value in
                              zip(json_data['loanManager'].keys(), coachAtts))/loanManagerSum
                headPhysio = sum(att_value * json_data['headPhysio'][att_name] for att_name, att_value in
                              zip(json_data['headPhysio'].keys(), coachAtts))/headPhysioSum
                headOfSportsScience = sum(att_value * json_data['headOfSportsScience'][att_name] for att_name, att_value in
                              zip(json_data['headOfSportsScience'].keys(), coachAtts))/headOfSportsScienceSum
                physio = sum(att_value * json_data['physio'][att_name] for att_name, att_value in
                              zip(json_data['physio'].keys(), coachAtts))/physioSum
                sportScientist = sum(att_value * json_data['sportScientist'][att_name] for att_name, att_value in
                              zip(json_data['sportScientist'].keys(), coachAtts))/sportScientistSum

                staff_dict = {
                    "info":row_data[0],
                    "name": row_data[1],
                    "age": row_data[2],
                    "personality": row_data[3],
                    "nationality": row_data[4],
                    "club":row_data[5],
                    "wage":row_data[6],
                    "atts":row_data[7:],
                    "rolesScores":[manager, assistantManager, headOfYouthDevelopment, goalkeepingCoachShotStopping,
                                    goalkeepingCoachHandlingandDistribution, fitnessCoachStrength, fitnessCoachQuickness, defendingTechnical,
                                    defendingTactical, possessionTechnical, possessionTactical, attackingTechnical, attackingTactical, headPerformanceAnalyst,
                                    performanceAnalyst, directorOfFootball, technicalDirector, chiefScout, scout, recruitmentAnalyst, loanManager, headPhysio,
                                    headOfSportsScience, physio, sportScientist],
                }
                data.append(staff_dict)

    new_workbook = Workbook()
    new_sheet = new_workbook.active

    # Başlık satırını ekleyin
    header = ["Info", "Name", "Age", "Personality", "Nationality", "Club", "Wage", "Manager", "Assistant Manager",
              "Head of Youth Development", "Goalkeeping Coach (Shot Stopping)",
              "Goalkeeping Coach (Handling and Distribution)", "Fitness Coach (Strength)", "Fitness Coach (Quickness)",
              "Defending (Technical)", "Defending (Tactical)", "Possession (Technical)", "Possession (Tactical)",
              "Attacking (Technical)", "Attacking (Tactical)", "Head Performance Analyst", "Performance Analyst",
              "Director of Football", "Technical Director", "Chief Scout", "Scout", "Recruitment Analyst",
              "Loan Manager", "Head Physio", "Head of Sports Science", "Physio", "Sport Scientist"]
    new_sheet.append(header)

    # Verileri ekleyin
    for staff_dict in data:
        row_data = [
                       staff_dict["info"],
                       staff_dict["name"],
                       staff_dict["age"],
                       staff_dict["personality"],
                       staff_dict["nationality"],
                       staff_dict["club"],
                       staff_dict["wage"],
                   ] + staff_dict["rolesScores"]

        new_sheet.append(row_data)

        for row in new_sheet.iter_rows(min_row=2, min_col=8):  # Start from the 3rd row and 8th column
            for cell in row:
                if cell.value is not None and float(cell.value) > 10 and float(cell.value) <= 12 :
                    cell.fill = yellow_fill
                elif cell.value is not None and float(cell.value) > 12:
                    cell.fill = green_fill

    # Dosyayı kaydedin
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dosyaları", "*.xlsx")])
    new_workbook.save(save_path)
# Denklemleri yazmak kaldı
def moneyballStaMentality():

    def calculate_moneyball_player_stats(json, player_stat_list):
        money_ball_player_stat_list = []
        for i, (key, value) in enumerate(json.items()):
            if i < len(player_stat_list):
                yeni_deger = value * player_stat_list[i]
                money_ball_player_stat_list.append(yeni_deger)
            else:
                money_ball_player_stat_list.append(None)
            if money_ball_player_stat_list[i] == "":
                money_ball_player_stat_list[i] = "0"
            money_ball_player_stat_list[i] = float(money_ball_player_stat_list[i])
        return money_ball_player_stat_list

    def calculateMoneyBallScore(input_list_name, data):
        if input_list_name is None or not data:
            return None

        operation = input_list_name

        if operation == "keeper":
            result = (data[3]+data[4]+data[16]+(data[36]+ data[37]+ data[38])*data[35])  / (data[0]/90)
            return result
        elif operation == "defender":
            result = (data[4] + 90*data[5]*data[21] + 90*data[7] + 90*data[9] + 90*data[20] +90*data[22] +90*data[23] +90*data[25] +90*data[31] + (90*data[43]*data[42]))  / (data[0])
            return result
        elif operation == "presser":
            result = (data[4] + data[7]*90 + data[9]*90 + data[13] + data[22]*90 + data[25]*90 + 90*data[31] + 90*data[32]+ 90*data[33] + (90*data[43]*data[42]) )  / (data[0])
            return result
        elif operation == "winger":
            result = (data[1]+ data[2] + data[4] + data[6]*90 + (90*data[11]*data[10]) + 10*data[14]*90)  / (data[0])
            return result
        elif operation == "creater":
            result = (data[1]+ 10*data[2] + data[4] + data[6]*90 + data[8]*90 + data[24]*90 + data[26]*90 + data[29]*90 + data[34]*90 ) / (data[0])
            return result
        elif operation == "finisher":
            result = (10*data[1]+ data[2] + data[4] + (90*data[41]*data[40]) + data[15]*90*data[41] + data[19]*90*data[41] )/ (data[0])
            return result
            for value in data:
                result *= value
            return result
        else:
            return None

    with open('playerSta.json', 'r') as json_file:
        json_data = json.load(json_file)

    keeperSum = sum(json_data['keeper'].values())
    defenderSum = sum(json_data['defender'].values())
    presserSum = sum(json_data['presser'].values())
    wingerSum = sum(json_data['winger'].values())
    createrSum = sum(json_data['creater'].values())
    finisherSum = sum(json_data['finisher'].values())

    file_path = filedialog.askopenfilename(filetypes=[("HTML Dosyaları", "*.html")])

    with open(file_path, 'r', encoding='utf-8') as file:
        html = file.read()
        soup = BeautifulSoup(html, 'html.parser')

        table = soup.find('table')

        headers = table.find_all('th')
        header_names = [header.text.strip() for header in headers]

        rows = table.find_all('tr')

        data = []

        for row in rows[1:]:
            cells = row.find_all('td')
            row_data = [cell.text.strip() for cell in cells]

            if len(row_data) == len(header_names):
                playerStats = []
                for stats in row_data[17:]:
                    try:
                        # Try to convert the stat to an integer
                        if "-" in stats:
                            stats_value = 0

                        stat_value = int(stats)
                    except ValueError:
                        # If it's not an integer, try converting to a float
                        try:
                            if "-" in stats:
                                stats_value = 0

                            stat_value = float(stats)
                        except ValueError:
                            # If it's neither int nor float, handle it as needed
                            stat_value = 0  # You can change this to another default value
                    playerStats.append(stat_value)

                #min to int
                if "," in row_data[13] :
                    row_data[13] = row_data[13].replace(",", '')
                    row_data[13] = float(row_data[13])
                sta = 0
                for i in range(len(row_data)):
                    item = row_data[i]
                    if isinstance(item, str) and "%" in item:
                        item = item.replace("%", "")
                        row_data[i] = float(item)
                    if isinstance(item, str) and "-" in item:
                        if item == "-":
                            item = item.replace("-", "0")
                            row_data[i] = float(item)
                    if isinstance(item, str) and "km" in item:
                        item = item.replace("km", "")
                        row_data[i] = float(item)
                    if i>13 :
                        row_data[i] = float(item)
                playerStatList = row_data[13:]
                print(playerStatList)

                keeperList = calculate_moneyball_player_stats(json_data["keeper"], playerStatList)
                defenderList = calculate_moneyball_player_stats(json_data["defender"], playerStatList)
                presserList = calculate_moneyball_player_stats(json_data["presser"], playerStatList)
                wingerList = calculate_moneyball_player_stats(json_data["winger"], playerStatList)
                createrList = calculate_moneyball_player_stats(json_data["creater"], playerStatList)
                finisherList = calculate_moneyball_player_stats(json_data["finisher"], playerStatList)

                keeper = calculateMoneyBallScore("keeper",playerStatList)
                defender = calculateMoneyBallScore("defender",playerStatList)
                presser = calculateMoneyBallScore("presser",playerStatList)
                winger = calculateMoneyBallScore("winger",playerStatList)
                creater = calculateMoneyBallScore("creater",playerStatList)
                finisher = calculateMoneyBallScore("finisher",playerStatList)

                row_data[14] = float(row_data[14])
                row_data[15] = float(row_data[15])
                row_data[16] = float(row_data[16])
                playerDict = {
                    "recomendation": row_data[0],
                    "info": row_data[1],
                    "name": row_data[2],
                    "age": row_data[3],
                    "nationality": row_data[4],
                    "personality": row_data[5],
                    "position":row_data[6],
                    "club": row_data[7],
                    "division":row_data[8],
                    "contract": row_data[9],
                    "wage":row_data[10],
                    "expires":row_data[11],
                    "value":row_data[12],
                    "mins":row_data[13],
                    "goals": row_data[14],
                    "assist": row_data[15],
                    "cleanSheets": row_data[16],
                    "stats":row_data[17:],
                    "rolesScores": [keeper, defender, presser,
                                    winger, creater, finisher],
                }
                data.append(playerDict)
                print(playerDict)

    new_workbook = Workbook()
    new_sheet = new_workbook.active

    header = ["Recomendation", "Info", "Name", "Age", "Nationality", "Personality", "Position", "Club", "Division", "Contract", "Wage",
              "Expires", "Value", "Mins", "Goals", "Assist","Clean Sheets",  "Average Rating", "aerial cha. attem. per 90",  "assists per 90",
              "blocks per 90",  "chances created per 90", "clearances per 90",  "cross comp. ratio",
              "cross attem. per 90",  "cross comp. ratio",  "distance covered per 90", "dribbles per 90",  "expected goals per shot ratio", "expected goals prevented",
              "expected goals prevented per 90", "expected save percentage",
              "goals per shot ratio", "headers won per 90", "headers won ratio", "interceptions per 90", "key headers per 90", "key passes per 90",
              "key tackles per 90", "open play key passes per 90", "pass attem. per 90", "pass comp. ratio",
              "pass comp. per 90", "posses. lost per 90", "posses. won per 90", "press attem. per 90", "press comp. per 90", "progress. passes per 90",
              "save ratio", "save held", "save parried", "save tipped", "shots on target per 90", "shots on target ratio",
              "shots per 90", "tackle comp. ratio", "tackle per 90",
              "Keeper", "Defender", "Presser", "Winger","Creater",
              "Finisher"]
    new_sheet.append(header)

    # Verileri ekleyin
    for playerDict in data:
        row_data = [
                       playerDict["recomendation"],
                       playerDict["info"],
                       playerDict["name"],
                       playerDict["age"],
                       playerDict["nationality"],
                       playerDict["personality"],
                       playerDict["position"],
                       playerDict["club"],
                       playerDict["division"],
                       playerDict["contract"],
                       playerDict["wage"],
                       playerDict["expires"],
                       playerDict["value"],
                       playerDict["mins"],
                       playerDict["goals"],
                       playerDict["assist"],
                       playerDict["cleanSheets"],
                   ] + playerDict["stats"]+ playerDict["rolesScores"]

        new_sheet.append(row_data)

    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Dosyaları", "*.xlsx")])
    new_workbook.save(save_path)

my_w.mainloop()